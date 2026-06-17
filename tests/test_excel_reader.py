from openpyxl import Workbook

from src.excel_reader import read_products
from src.excel_reader import validate_required_columns

def test_read_products(tmp_path):
    filepath = tmp_path / "products.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article_1",
        "Title_1",
        "Size_1",
        "Price_1",
        "Currency_1"
    ])

    sheet.append([
        "2x7x01",
        "ICEBERG white",
        "40 (x)",
        290,
        "usd"
    ])

    workbook.save(filepath)

    profile_config = {
        "columns": {
            "article": "Article_1",
            "name": "Title_1",
            "size": "Size_1",
            "price": "Price_1",
            "currency": "Currency_1"
        },
        "product_key_fields": [
            "article",
            "size"
        ],
        "label_item_fields": [
            "article",
            "name",
            "size"
        ]
    }

    result = read_products(
        filepath,
        profile_config
    )

    assert result == [
        {
            "_row_index": 2,
            "article": "2x7x01",
            "name": "ICEBERG white",
            "size": "40 (x)",
            "price": 290,
            "currency": "usd"
        }
    ]


def test_read_products_without_size_column(tmp_path):
    filepath = tmp_path / "bags.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article",
        "Name",
        "Price",
        "Currency"
    ])

    sheet.append([
        "BG-001",
        "Leather Bag",
        149,
        "USD"
    ])

    workbook.save(filepath)

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    result = read_products(
        filepath,
        profile_config
    )

    assert result == [
        {
            "_row_index": 2,
            "article": "BG-001",
            "name": "Leather Bag",
            "price": 149,
            "currency": "USD"
        }
    ]


def test_read_products_ignores_extra_columns(tmp_path):
    filepath = tmp_path / "products.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article",
        "Name",
        "Price",
        "Currency",
        "Color",
        "Country"
    ])

    sheet.append([
        "BG-001",
        "Leather Bag",
        149,
        "USD",
        "Black",
        "Italy"
    ])

    workbook.save(filepath)

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    result = read_products(
        filepath,
        profile_config
    )

    assert result == [
        {
            "_row_index": 2,
            "article": "BG-001",
            "name": "Leather Bag",
            "price": 149,
            "currency": "USD"
        }
    ]


def test_read_products_raises_exception_when_required_column_is_missing(tmp_path):
    filepath = tmp_path / "products.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article",
        "Name",
        "Currency"
    ])

    sheet.append([
        "BG-001",
        "Leather Bag",
        "USD"
    ])

    workbook.save(filepath)

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    try:
        read_products(
            filepath,
            profile_config
        )

        assert False
    except ValueError as error:
        assert str(error) == "Required column 'price' not found in Excel file"


def test_read_products_does_not_require_barcode_column(tmp_path):
    filepath = tmp_path / "products.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article",
        "Name",
        "Price",
        "Currency"
    ])

    sheet.append([
        "BG-001",
        "Leather Bag",
        149,
        "USD"
    ])

    workbook.save(filepath)

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency",
            "barcode": "Barcode"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    result = read_products(
        filepath,
        profile_config
    )

    assert result == [
        {
            "_row_index": 2,
            "article": "BG-001",
            "name": "Leather Bag",
            "price": 149,
            "currency": "USD"
        }
    ]


def test_validate_required_columns_raises_exception_when_required_column_is_missing():
    headers = [
        "Article",
        "Name",
        "Currency"
    ]

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    try:
        validate_required_columns(
            headers,
            profile_config
        )

        assert False
    except ValueError as error:
        assert str(error) == "Required column 'price' not found in Excel file"


def test_validate_required_columns_accepts_existing_required_columns():
    headers = [
        "Article",
        "Name",
        "Price",
        "Currency"
    ]

    profile_config = {
        "columns": {
            "article": "Article",
            "name": "Name",
            "price": "Price",
            "currency": "Currency"
        },
        "product_key_fields": [
            "article"
        ],
        "label_item_fields": [
            "article",
            "name"
        ]
    }

    validate_required_columns(
        headers,
        profile_config
    )