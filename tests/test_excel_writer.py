from openpyxl import Workbook
from openpyxl import load_workbook

from src.excel_writer import write_products_with_barcodes


def test_write_products_with_new_barcode_column(tmp_path):
    input_filepath = tmp_path / "input.xlsx"
    output_filepath = tmp_path / "output.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article_x",
        "Size_y"
    ])

    sheet.append([
        "2x7x01",
        "40(p)"
    ])

    workbook.save(input_filepath)

    barcodes_by_row = {
        2: "2966150061736"
    }

    column_mapping = {
        "article": "Article_x",
        "size": "Size_y",
        "barcode": "Barcode_z"
    }

    write_products_with_barcodes(
        input_filepath,
        output_filepath,
        barcodes_by_row,
        column_mapping
    )

    result_workbook = load_workbook(output_filepath)
    result_sheet = result_workbook.active

    assert result_sheet.cell(1, 3).value == "Barcode_z"
    assert result_sheet.cell(2, 3).value == "2966150061736"


def test_write_products_with_existing_barcode_column(tmp_path):
    input_filepath = tmp_path / "input.xlsx"
    output_filepath = tmp_path / "output.xlsx"

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Article_x",
        "Size_y",
        "Barcode_z"
    ])

    sheet.append([
        "2x7x01",
        "40(p)",
        ""
    ])

    workbook.save(input_filepath)

    barcodes_by_row = {
        2: "2966150061736"
    }

    column_mapping = {
        "article": "Article_x",
        "size": "Size_y",
        "barcode": "Barcode_z"
    }

    write_products_with_barcodes(
        input_filepath,
        output_filepath,
        barcodes_by_row,
        column_mapping
    )

    result_workbook = load_workbook(output_filepath)
    result_sheet = result_workbook.active

    assert result_sheet.cell(2, 3).value == "2966150061736"