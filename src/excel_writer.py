from pathlib import Path

from openpyxl import load_workbook


def write_products_with_barcodes(
    input_filepath,
    output_filepath,
    barcodes_by_row,
    column_mapping
):
    workbook = load_workbook(input_filepath)

    sheet = workbook.active

    headers = []

    for cell in sheet[1]:
        headers.append(cell.value)

    barcode_column_name = column_mapping.get("barcode") or "barcode"
    barcode_column_index = len(headers) + 1

    if barcode_column_name in headers:
        barcode_column_index = headers.index(barcode_column_name) + 1
    else:
        sheet.cell(
            row=1,
            column=barcode_column_index,
            value=barcode_column_name
        )

    for row_index, barcode in barcodes_by_row.items():
        sheet.cell(
            row=row_index,
            column=barcode_column_index,
            value=barcode
        )

    output_path = Path(output_filepath)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(output_filepath)