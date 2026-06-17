from openpyxl import load_workbook

REQUIRED_COLUMN_NAMES = [
    "price",
    "currency"
]


def build_required_column_names(profile_config):
    result = []

    result.extend(profile_config["product_key_fields"])
    result.extend(profile_config["label_item_fields"])
    result.extend(REQUIRED_COLUMN_NAMES)

    return set(result)


def validate_required_columns(
    headers,
    profile_config
):
    column_mapping = profile_config["columns"]
    required_column_names = build_required_column_names(profile_config)

    for internal_name in required_column_names:
        excel_column_name = column_mapping[internal_name]

        if excel_column_name not in headers:
            raise ValueError(
                f"Required column '{internal_name}' not found in Excel file"
            )
        

def read_products(
    filepath,
    profile_config
):
    column_mapping = profile_config["columns"]
    workbook = load_workbook(filepath)

    sheet = workbook.active

    rows = list(
        sheet.iter_rows(values_only=True)
    )

    headers = rows[0]

    validate_required_columns(
        headers,
        profile_config
    )

    header_indexes = {}

    for internal_name, excel_column_name in column_mapping.items():
        if excel_column_name not in headers:
            continue

        header_indexes[internal_name] = headers.index(
            excel_column_name
        )

    result = []

    for row_index, row in enumerate(rows[1:], start=2):
        is_empty = True

        for value in row:
            if value not in (None, ""):
                is_empty = False
                break

        if is_empty:
            continue

        item = {
            "_row_index": row_index
        }

        for internal_name, index in header_indexes.items():
            item[internal_name] = row[index]

        result.append(item)

    return result